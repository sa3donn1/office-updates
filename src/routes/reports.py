from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from datetime import datetime, timedelta
from src.routes.auth import login_required
from src.models.employee import Employee
from src.models.client import Client
from src.models.task import Task
from src.models.user import db
from sqlalchemy import func, and_, or_
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
#from weasyprint import HTML, CSS
#from weasyprint.text.fonts import FontConfiguration

reports_bp = Blueprint("reports", __name__)

@reports_bp.route("/")
@login_required
def index():
    """صفحة التقارير الرئيسية"""
    return render_template("reports/index.html")

@reports_bp.route("/employee_productivity")
@login_required
def employee_productivity():
    """تقرير إنتاجية الموظفين"""
    # جلب جميع الموظفين مع إحصائياتهم
    employees_data = []
    employees = Employee.query.all()
    
    for employee in employees:
        # عدد المهام المكتملة
        completed_tasks = Task.query.filter_by(employee_id=employee.id, is_done=True).count()
        
        # عدد العملاء المسؤول عنهم
        clients_count = Client.query.filter_by(employee_id=employee.id).count()
        
        # إجمالي المهام
        total_tasks = Task.query.filter_by(employee_id=employee.id).count()
        
        # نسبة الإنجاز
        completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        employees_data.append({
            'employee': employee,
            'completed_tasks': completed_tasks,
            'total_tasks': total_tasks,
            'clients_count': clients_count,
            'completion_rate': round(completion_rate, 2)
        })
    
    # ترتيب حسب عدد المهام المكتملة
    employees_data.sort(key=lambda x: x['completed_tasks'], reverse=True)
    
    return render_template("reports/employee_productivity.html", employees_data=employees_data)

@reports_bp.route("/tasks_by_employee")
@login_required
def tasks_by_employee():
    """تقرير المهام حسب الموظف"""
    selected_employee_id = request.args.get('employee_id', type=int)
    employees = Employee.query.all()
    
    tasks_data = []
    selected_employee = None
    
    if selected_employee_id:
        selected_employee = Employee.query.get(selected_employee_id)
        if selected_employee:
            tasks = Task.query.filter_by(employee_id=selected_employee_id).all()
            
            # تصنيف المهام حسب الحالة
            completed_tasks = [task for task in tasks if task.is_done]
            pending_tasks = [task for task in tasks if not task.is_done]
            
            tasks_data = {
                'completed': completed_tasks,
                'pending': pending_tasks,
                'total': len(tasks)
            }
    
    return render_template("reports/tasks_by_employee.html", 
                         employees=employees, 
                         selected_employee=selected_employee,
                         tasks_data=tasks_data)

@reports_bp.route("/overdue_tasks")
@login_required
def overdue_tasks():
    """تقرير المهام المتأخرة"""
    # في هذا التطبيق، لا يوجد حقل due_date في نموذج Task
    # لذلك سنعتبر المهام التي تم إنشاؤها منذ أكثر من 30 يوماً ولم تكتمل بعد كمهام متأخرة
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    
    overdue_tasks = Task.query.filter(
        and_(
            Task.is_done == False,
            Task.created_at < thirty_days_ago
        )
    ).all()
    
    return render_template("reports/overdue_tasks.html", overdue_tasks=overdue_tasks)

@reports_bp.route("/dashboard")
@login_required
def dashboard():
    """لوحة التحكم مع الإحصائيات والرسوم البيانية"""
    # الحصول على الفترة الزمنية من المعاملات
    period = request.args.get('period', 'all')  # all, today, week, month
    
    # تحديد التاريخ بناءً على الفترة المحددة
    date_filter = None
    if period == 'today':
        date_filter = datetime.utcnow().date()
    elif period == 'week':
        date_filter = datetime.utcnow() - timedelta(days=7)
    elif period == 'month':
        date_filter = datetime.utcnow() - timedelta(days=30)
    
    # الإحصائيات السريعة
    total_clients = Client.query.count()
    
    if date_filter:
        if period == 'today':
            total_tasks = Task.query.filter(func.date(Task.created_at) == date_filter).count()
            completed_tasks = Task.query.filter(
                and_(func.date(Task.created_at) == date_filter, Task.is_done == True)
            ).count()
        else:
            total_tasks = Task.query.filter(Task.created_at >= date_filter).count()
            completed_tasks = Task.query.filter(
                and_(Task.created_at >= date_filter, Task.is_done == True)
            ).count()
    else:
        total_tasks = Task.query.count()
        completed_tasks = Task.query.filter_by(is_done=True).count()
    
    # المهام المتأخرة (أكثر من 30 يوماً)
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    overdue_tasks = Task.query.filter(
        and_(Task.is_done == False, Task.created_at < thirty_days_ago)
    ).count()
    
    # بيانات الرسم البياني الدائري - توزيع المهام حسب الحالة
    pending_tasks = total_tasks - completed_tasks
    pie_data = {
        'completed': completed_tasks,
        'pending': pending_tasks,
        'overdue': overdue_tasks
    }
    
    # بيانات الرسم البياني الشريطي - المهام المكتملة لكل موظف
    employees = Employee.query.all()
    bar_data = []
    
    for employee in employees:
        if date_filter:
            if period == 'today':
                emp_completed = Task.query.filter(
                    and_(
                        Task.employee_id == employee.id,
                        Task.is_done == True,
                        func.date(Task.created_at) == date_filter
                    )
                ).count()
            else:
                emp_completed = Task.query.filter(
                    and_(
                        Task.employee_id == employee.id,
                        Task.is_done == True,
                        Task.created_at >= date_filter
                    )
                ).count()
        else:
            emp_completed = Task.query.filter_by(employee_id=employee.id, is_done=True).count()
        
        bar_data.append({
            'name': employee.name,
            'completed_tasks': emp_completed
        })
    
    # ترتيب حسب عدد المهام المكتملة
    bar_data.sort(key=lambda x: x['completed_tasks'], reverse=True)
    
    stats = {
        'total_clients': total_clients,
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'overdue_tasks': overdue_tasks
    }
    
    return render_template("reports/dashboard.html", 
                         stats=stats, 
                         pie_data=pie_data, 
                         bar_data=bar_data,
                         current_period=period)

@reports_bp.route("/export_excel/<report_type>")
@login_required
def export_excel(report_type):
    """تصدير التقارير إلى Excel"""
    wb = Workbook()
    ws = wb.active
    
    # إعداد الخط والتنسيق
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    if report_type == "employee_productivity":
        ws.title = "تقرير إنتاجية الموظفين"
        
        # العناوين
        headers = ["اسم الموظف", "المهام المكتملة", "إجمالي المهام", "عدد العملاء", "نسبة الإنجاز %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # البيانات
        employees = Employee.query.all()
        for row, employee in enumerate(employees, 2):
            completed_tasks = Task.query.filter_by(employee_id=employee.id, is_done=True).count()
            total_tasks = Task.query.filter_by(employee_id=employee.id).count()
            clients_count = Client.query.filter_by(employee_id=employee.id).count()
            completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
            
            ws.cell(row=row, column=1, value=employee.name)
            ws.cell(row=row, column=2, value=completed_tasks)
            ws.cell(row=row, column=3, value=total_tasks)
            ws.cell(row=row, column=4, value=clients_count)
            ws.cell(row=row, column=5, value=round(completion_rate, 2))
    
    elif report_type == "overdue_tasks":
        ws.title = "المهام المتأخرة"
        
        # العناوين
        headers = ["اسم الشركة", "نوع الخدمة", "الموظف المسؤول", "تاريخ الإنشاء"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # البيانات
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        overdue_tasks = Task.query.filter(
            and_(Task.is_done == False, Task.created_at < thirty_days_ago)
        ).all()
        
        for row, task in enumerate(overdue_tasks, 2):
            ws.cell(row=row, column=1, value=task.company_name)
            ws.cell(row=row, column=2, value=task.service_type)
            ws.cell(row=row, column=3, value=task.responsible_employee.name if task.responsible_employee else "غير محدد")
            ws.cell(row=row, column=4, value=task.created_at.strftime("%Y-%m-%d") if task.created_at else "")
    
    # حفظ الملف في الذاكرة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports_bp.route("/export_pdf/<report_type>")
@login_required
def export_pdf(report_type):
    """تصدير التقارير إلى PDF"""
    if report_type == "employee_productivity":
        # جلب البيانات
        employees_data = []
        employees = Employee.query.all()
        
        for employee in employees:
            completed_tasks = Task.query.filter_by(employee_id=employee.id, is_done=True).count()
            clients_count = Client.query.filter_by(employee_id=employee.id).count()
            total_tasks = Task.query.filter_by(employee_id=employee.id).count()
            completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
            
            employees_data.append({
                'employee': employee,
                'completed_tasks': completed_tasks,
                'total_tasks': total_tasks,
                'clients_count': clients_count,
                'completion_rate': round(completion_rate, 2)
            })
        
        employees_data.sort(key=lambda x: x['completed_tasks'], reverse=True)
        
        # إنشاء HTML للتقرير
        html_content = render_template("reports/pdf_employee_productivity.html", 
                                     employees_data=employees_data,
                                     generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"))
    
    elif report_type == "overdue_tasks":
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        overdue_tasks = Task.query.filter(
            and_(Task.is_done == False, Task.created_at < thirty_days_ago)
        ).all()
        
        html_content = render_template("reports/pdf_overdue_tasks.html", 
                                     overdue_tasks=overdue_tasks,
                                     generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"))
    
   