from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, session, send_file
from src.routes.auth import login_required
from src.models.attendance import Attendance
from src.models.employee import Employee
from src.models.user import db
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

attendance_bp = Blueprint("attendance", __name__)

# ============ موظف - تسجيل الحضور والانصراف ============

@attendance_bp.route("/check-in/<int:employee_id>", methods=["POST"])
@login_required
def check_in(employee_id):
    """تسجيل حضور الموظف"""
    try:
        # التحقق من أن الموظف يقوم بتسجيل حضوره فقط
        if session.get("employee_id") != employee_id and "admin_id" not in session:
            return jsonify({"success": False, "message": "ليس لديك صلاحية للقيام بهذا الإجراء"}), 403
        
        employee = Employee.query.get(employee_id)
        if not employee:
            return jsonify({"success": False, "message": "الموظف غير موجود"}), 404
        
        today = date.today()
        
        # البحث عن سجل اليوم
        attendance = Attendance.query.filter_by(
            employee_id=employee_id, 
            date=today
        ).first()
        
        if attendance:
            if attendance.check_in:
                return jsonify({
                    "success": False, 
                    "message": "تم تسجيل الحضور بالفعل اليوم",
                    "check_in_time": attendance.check_in.strftime("%H:%M:%S")
                }), 400
        else:
            # إنشاء سجل جديد
            attendance = Attendance(
                employee_id=employee_id,
                date=today,
                check_in=datetime.now()
            )
            db.session.add(attendance)
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "تم تسجيل الحضور بنجاح",
            "check_in_time": datetime.now().strftime("%H:%M:%S")
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"حدث خطأ: {str(e)}"}), 500


@attendance_bp.route("/check-out/<int:employee_id>", methods=["POST"])
@login_required
def check_out(employee_id):
    """تسجيل انصراف الموظف"""
    try:
        # التحقق من أن الموظف يقوم بتسجيل انصرافه فقط
        if session.get("employee_id") != employee_id and "admin_id" not in session:
            return jsonify({"success": False, "message": "ليس لديك صلاحية للقيام بهذا الإجراء"}), 403
        
        employee = Employee.query.get(employee_id)
        if not employee:
            return jsonify({"success": False, "message": "الموظف غير موجود"}), 404
        
        today = date.today()
        
        # البحث عن سجل اليوم
        attendance = Attendance.query.filter_by(
            employee_id=employee_id,
            date=today
        ).first()
        
        if not attendance:
            return jsonify({
                "success": False,
                "message": "لم يتم تسجيل حضور اليوم. يرجى تسجيل الحضور أولاً"
            }), 400
        
        if not attendance.check_in:
            return jsonify({
                "success": False,
                "message": "لم يتم تسجيل حضور. يرجى تسجيل الحضور أولاً"
            }), 400
        
        if attendance.check_out:
            return jsonify({
                "success": False,
                "message": "تم تسجيل الانصراف بالفعل",
                "check_out_time": attendance.check_out.strftime("%H:%M:%S")
            }), 400
        
        # تسجيل الانصراف
        attendance.check_out = datetime.now()
        db.session.commit()
        
        working_hours = attendance.get_working_hours()
        
        return jsonify({
            "success": True,
            "message": "تم تسجيل الانصراف بنجاح",
            "check_out_time": datetime.now().strftime("%H:%M:%S"),
            "working_hours": working_hours
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"حدث خطأ: {str(e)}"}), 500


# ============ أدمن - عرض السجلات والتحكم ============

@attendance_bp.route("/admin/list")
@login_required
def admin_list():
    """عرض قائمة الحضور والانصراف للأدمن"""
    if "admin_id" not in session:
        flash("ليس لديك صلاحية الوصول إلى هذه الصفحة", "danger")
        return redirect(url_for("auth.login"))
    
    page = request.args.get("page", 1, type=int)
    employee_filter = request.args.get("employee", "", type=str)
    date_filter = request.args.get("date", "", type=str)
    
    query = Attendance.query
    
    # تطبيق الفلاتر
    if employee_filter:
        query = query.join(Employee).filter(
            Employee.name.ilike(f"%{employee_filter}%")
        )
    
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
            query = query.filter(Attendance.date == filter_date)
        except ValueError:
            pass
    
    # ترتيب حسب التاريخ الأحدث أولاً
    attendance_records = query.order_by(Attendance.date.desc()).paginate(page=page, per_page=20)
    
    employees = Employee.query.all()
    
    return render_template(
        "attendance/admin_list.html",
        attendance_records=attendance_records,
        employees=employees,
        current_employee=employee_filter,
        current_date=date_filter
    )


@attendance_bp.route("/admin/export", methods=["GET"])
@login_required
def export_excel():
    """تصدير سجلات الحضور إلى ملف Excel"""
    if "admin_id" not in session:
        return jsonify({"success": False, "message": "ليس لديك صلاحية"}), 403
    
    try:
        # الحصول على الفلاتر من الاستعلام
        employee_filter = request.args.get("employee", "", type=str)
        date_filter = request.args.get("date", "", type=str)
        
        query = Attendance.query
        
        if employee_filter:
            query = query.join(Employee).filter(
                Employee.name.ilike(f"%{employee_filter}%")
            )
        
        if date_filter:
            try:
                filter_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
                query = query.filter(Attendance.date == filter_date)
            except ValueError:
                pass
        
        records = query.order_by(Attendance.date.desc()).all()
        
        # إنشاء كتاب Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "الحضور والانصراف"
        
        # تعيين عرض الأعمدة
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 25
        
        # رؤوس الجداول
        headers = ["#", "اسم الموظف", "التاريخ", "وقت الحضور", "وقت الانصراف", "ساعات العمل", "ملاحظات"]
        
        # إضافة رؤوس الأعمدة مع التنسيق
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # إضافة البيانات
        for row_num, record in enumerate(records, 2):
            sheet.cell(row=row_num, column=1).value = row_num - 1
            sheet.cell(row=row_num, column=2).value = record.employee.name
            sheet.cell(row=row_num, column=3).value = record.date.strftime("%Y-%m-%d")
            sheet.cell(row=row_num, column=4).value = record.check_in.strftime("%H:%M:%S") if record.check_in else "-"
            sheet.cell(row=row_num, column=5).value = record.check_out.strftime("%H:%M:%S") if record.check_out else "-"
            sheet.cell(row=row_num, column=6).value = record.get_working_hours() if record.get_working_hours() else "-"
            sheet.cell(row=row_num, column=7).value = record.notes or "-"
            
            # توسيط البيانات
            for col in range(1, 8):
                sheet.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")
        
        # حفظ الكتاب في BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'attendance_report_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({"success": False, "message": f"حدث خطأ: {str(e)}"}), 500


@attendance_bp.route("/admin/edit/<int:attendance_id>", methods=["GET", "POST"])
@login_required
def edit_attendance(attendance_id):
    """تعديل سجل حضور (للأدمن فقط)"""
    if "admin_id" not in session:
        flash("ليس لديك صلاحية الوصول إلى هذه الصفحة", "danger")
        return redirect(url_for("auth.login"))
    
    attendance = Attendance.query.get_or_404(attendance_id)
    
    if request.method == "POST":
        try:
            check_in_time = request.form.get("check_in")
            check_out_time = request.form.get("check_out")
            notes = request.form.get("notes")
            
            if check_in_time:
                attendance.check_in = datetime.strptime(check_in_time, "%H:%M")
                attendance.check_in = attendance.check_in.replace(
                    year=attendance.date.year,
                    month=attendance.date.month,
                    day=attendance.date.day
                )
            
            if check_out_time:
                attendance.check_out = datetime.strptime(check_out_time, "%H:%M")
                attendance.check_out = attendance.check_out.replace(
                    year=attendance.date.year,
                    month=attendance.date.month,
                    day=attendance.date.day
                )
            
            attendance.notes = notes
            db.session.commit()
            
            flash("تم تحديث السجل بنجاح", "success")
            return redirect(url_for("attendance.admin_list"))
        
        except Exception as e:
            db.session.rollback()
            flash(f"حدث خطأ: {str(e)}", "danger")
    
    return render_template("attendance/edit.html", attendance=attendance)


@attendance_bp.route("/admin/delete/<int:attendance_id>", methods=["POST"])
@login_required
def delete_attendance(attendance_id):
    """حذف سجل حضور (للأدمن فقط)"""
    if "admin_id" not in session:
        return jsonify({"success": False, "message": "ليس لديك صلاحية"}), 403
    
    try:
        attendance = Attendance.query.get_or_404(attendance_id)
        employee_name = attendance.employee.name
        
        db.session.delete(attendance)
        db.session.commit()
        
        flash(f"تم حذف سجل حضور {employee_name} بنجاح", "success")
        return redirect(url_for("attendance.admin_list"))
        
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ: {str(e)}", "danger")
        return redirect(url_for("attendance.admin_list"))


@attendance_bp.route("/status/<int:employee_id>")
@login_required
def get_status(employee_id):
    """الحصول على حالة الحضور والانصراف لليوم"""
    try:
        if session.get("employee_id") != employee_id and "admin_id" not in session:
            return jsonify({"success": False, "message": "ليس لديك صلاحية"}), 403
        
        today = date.today()
        attendance = Attendance.query.filter_by(
            employee_id=employee_id,
            date=today
        ).first()
        
        status = {
            "checked_in": False,
            "checked_out": False,
            "check_in_time": None,
            "check_out_time": None,
            "working_hours": None
        }
        
        if attendance:
            status["checked_in"] = attendance.check_in is not None
            status["checked_out"] = attendance.check_out is not None
            if attendance.check_in:
                status["check_in_time"] = attendance.check_in.strftime("%H:%M:%S")
            if attendance.check_out:
                status["check_out_time"] = attendance.check_out.strftime("%H:%M:%S")
            status["working_hours"] = attendance.get_working_hours()
        
        return jsonify({"success": True, "data": status}), 200
        
    except Exception as e:
        return jsonify({"success": False, "message": f"حدث خطأ: {str(e)}"}), 500
